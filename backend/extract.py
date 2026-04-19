"""
Extracción PDF → Excel con Claude Vision — FacturAI Backend
Prompt actualizado con reglas contables completas.
"""
import re
import io
import base64
import logging
from datetime import datetime, date as date_type
from typing import Any, Dict, List, Optional, Tuple

import fitz  # PyMuPDF
import anthropic
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────
# UTILIDADES NUMÉRICAS Y DE FECHA
# ─────────────────────────────────────────────

def safe_float(value: Any) -> Optional[float]:
    """Parsea importes en cualquier formato europeo/americano."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        v = round(float(value), 2)
        return v if 0 < v < 1_000_000 else None
    text = str(value).strip()
    if not text:
        return None
    text = re.sub(r'[€$£\s]', '', text)
    text = re.sub(r'(EUR|eur|Eur)', '', text)
    if re.match(r'^\d{1,3}(\.\d{3})+(,\d+)?$', text):
        text = text.replace('.', '').replace(',', '.')
    elif re.match(r'^\d+,\d{1,2}$', text):
        text = text.replace(',', '.')
    elif re.match(r'^\d{1,3}(,\d{3})+(\.\d+)?$', text):
        text = text.replace(',', '')
    else:
        text = re.sub(r'[^\d.]', '', text)
    try:
        result = round(float(text), 2)
        return result if 0 < result < 1_000_000 else None
    except Exception:
        return None


def parse_date_to_iso(date_text: Any) -> Optional[str]:
    if not date_text:
        return None
    raw = str(date_text).strip()
    patterns = [
        "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y",
        "%d.%m.%Y", "%Y/%m/%d", "%d-%m-%y", "%d/%m/%y",
    ]
    for fmt in patterns:
        try:
            return datetime.strptime(raw, fmt).strftime("%d-%m-%Y")
        except Exception:
            pass
    match = re.search(r"(\d{2})[\/\-.](\d{2})[\/\-.](\d{4})", raw)
    if match:
        d, m, y = match.groups()
        try:
            return datetime(int(y), int(m), int(d)).strftime("%d-%m-%Y")
        except Exception:
            return None
    return None


def day_of_year(date_iso: str) -> Optional[int]:
    """Calcula el día del año (1-366) a partir de DD-MM-YYYY."""
    try:
        parts = date_iso.split("-")
        d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
        return date_type(y, m, d).timetuple().tm_yday
    except Exception:
        return None


def date_to_sortable(date_iso: str) -> str:
    """Convierte DD-MM-YYYY a YYYY-MM-DD para ordenación."""
    try:
        parts = date_iso.split("-")
        if len(parts) == 3 and len(parts[2]) == 4:
            return f"{parts[2]}-{parts[1]}-{parts[0]}"
    except Exception:
        pass
    return date_iso


def normalize_estado(value: Any) -> str:
    text = str(value or "").strip().upper()
    mapping = {
        "COMPLETA": "COMPLETA", "OK": "COMPLETA", "COMPLETE": "COMPLETA",
        "VERIFICAR_DATOS": "VERIFICAR_DATOS", "VERIFICAR": "VERIFICAR_DATOS",
        "REVISAR": "VERIFICAR_DATOS", "PENDIENTE_REVISION": "PENDIENTE_REVISION",
        "PENDIENTE": "PENDIENTE_REVISION", "NO_PROCESABLE": "PENDIENTE_REVISION",
    }
    return mapping.get(text, "PENDIENTE_REVISION")


def build_fallback_row(page_num: int, reason: str) -> Dict[str, Any]:
    return {
        "numero_factura": None, "pagina": page_num,
        "fecha_literal": None, "fecha_iso": None,
        "total_eur": None, "iva_pct": 10,
        "base_eur": None, "cuota_eur": None,
        "estado": "PENDIENTE_REVISION",
        "observaciones": reason[:250] if reason else "No procesable",
        "concepto": "venta a clientes varios",
        "tipo_impositivo_iva": 10,
    }


def normalize_row(row: Dict[str, Any], page_num: int) -> Dict[str, Any]:
    total = safe_float(row.get("total_eur"))
    # Base = Total / 1.10 siempre (regla fija del prompt)
    base  = round(total / 1.10, 2) if total is not None else None
    cuota = round(base * 0.10, 2)  if base  is not None else None

    fecha_iso = parse_date_to_iso(row.get("fecha_iso") or row.get("fecha_literal"))

    estado = normalize_estado(row.get("estado"))
    obs    = str(row.get("observaciones") or "OK").strip()

    # Ajustar estado según datos disponibles
    if fecha_iso and total is not None:
        estado = "COMPLETA"
    elif fecha_iso is None and total is None:
        if estado != "PENDIENTE_REVISION":
            estado = "PENDIENTE_REVISION"
    else:
        if estado == "COMPLETA":
            estado = "VERIFICAR_DATOS"

    return {
        "numero_factura":     row.get("numero_factura"),
        "pagina":             row.get("pagina") if row.get("pagina") is not None else page_num,
        "fecha_literal":      row.get("fecha_literal"),
        "fecha_iso":          fecha_iso,
        "total_eur":          total,
        "iva_pct":            10,
        "base_eur":           base,
        "cuota_eur":          cuota,
        "estado":             estado,
        "observaciones":      obs,
        "concepto":           "venta a clientes varios",
        "tipo_impositivo_iva": 10,
    }


# ─────────────────────────────────────────────
# NUMERACIÓN Z-{DÍA DEL AÑO}
# ─────────────────────────────────────────────

def sort_and_renumber_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Ordena por Fecha ISO ascendente (sin fecha al final).
    Asigna Z-{día del año} a filas fechadas, con sufijo a/b/c si coinciden.
    Las no fechadas continúan el correlativo desde el último número asignado.
    """
    dated   = [r for r in rows if r.get("fecha_iso")]
    undated = [r for r in rows if not r.get("fecha_iso")]

    # Ordenar fechadas de más antigua a más reciente
    dated.sort(key=lambda r: date_to_sortable(r["fecha_iso"]))

    # Contar cuántas facturas caen en cada día del año
    day_count: Dict[int, int] = {}
    for r in dated:
        doy = day_of_year(r["fecha_iso"])
        if doy:
            day_count[doy] = day_count.get(doy, 0) + 1

    # Asignar número de factura
    day_used: Dict[int, int] = {}
    for r in dated:
        doy = day_of_year(r["fecha_iso"])
        if doy is None:
            r["numero_factura"] = "Z-?"
        elif day_count[doy] == 1:
            r["numero_factura"] = f"Z-{doy}"
        else:
            day_used[doy] = day_used.get(doy, 0) + 1
            suffix = chr(ord('a') + day_used[doy] - 1)
            r["numero_factura"] = f"Z-{doy}{suffix}"

    # Facturas sin fecha: continuar correlativo desde el último día usado
    last_doy = max(day_count.keys()) if day_count else 0
    for i, r in enumerate(undated, start=1):
        r["numero_factura"] = f"Z-{last_doy + i}"

    return dated + undated


# ─────────────────────────────────────────────
# PROMPT CLAUDE VISION
# ─────────────────────────────────────────────

SYSTEM_PROMPT = """Eres un extractor contable especializado en facturas, tickets y comprobantes españoles en PDF.

TAREA: Analizar esta página y extraer los datos de la factura o ticket visible.

REGLAS DE EXTRACCIÓN:

FECHA:
- Busca campos: Fecha, Date, F., Emitido, Emisión, Fcha. o patrones DD/MM/YYYY, DD-MM-YYYY, YYYY-MM-DD
- Conserva el texto original en fecha_literal aunque sea ambiguo
- Normaliza a DD-MM-YYYY en fecha_iso solo si puedes hacerlo con seguridad
- NO uses fechas de vencimiento, entrega o pago posterior
- Si no puedes determinar el año con certeza, fecha_iso = null

TOTAL:
- El total es el importe FINAL pagado (incluye IVA)
- Busca: Total, TOTAL, Importe total, A pagar, Total a pagar
- NO confundas con: Base Imponible, Subtotal, Descuento, IVA por separado
- Los números en España: punto = miles, coma = decimal → 1.234,56 € = 1234.56
- Devuelve total_eur como número decimal SIN símbolo: 1234.56

IVA:
- Devuelve solo el porcentaje numérico: 10, 21, 4 (sin símbolo %)
- Si no está visible, usa 10 por defecto

ESTADO:
- "COMPLETA": tienes fecha_iso Y total_eur con confianza
- "VERIFICAR_DATOS": falta uno de los dos o calidad visual baja
- "PENDIENTE_REVISION": página vacía, ilegible, no es factura/ticket

RESPUESTA: JSON puro sin markdown, sin explicaciones.

{
  "pagina": <entero>,
  "fecha_literal": "<texto original o null>",
  "fecha_iso": "<DD-MM-YYYY o null>",
  "total_eur": <decimal sin símbolo o null>,
  "iva_pct": <número o null>,
  "estado": "COMPLETA" | "VERIFICAR_DATOS" | "PENDIENTE_REVISION",
  "observaciones": "<máx 20 palabras>"
}"""


async def extract_invoice_from_page(
    client: anthropic.AsyncAnthropic,
    page_b64: str,
    page_num: int,
    model: str = "claude-opus-4-5",
) -> Dict[str, Any]:
    user_prompt = (
        f"Página {page_num}. "
        "Extrae fecha de emisión y total final con IVA incluido. "
        "Si ves importes, DEBES extraer el total. "
        "Responde SOLO con el JSON."
    )
    try:
        response = await client.messages.create(
            model=model,
            max_tokens=512,
            system=SYSTEM_PROMPT,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": "image/png", "data": page_b64}},
                    {"type": "text", "text": user_prompt},
                ],
            }],
        )
        raw = response.content[0].text.strip()
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw).strip()
        if not raw:
            return build_fallback_row(page_num, "Sin respuesta del modelo")
        parsed = json.loads(raw)
        if not isinstance(parsed, dict):
            return build_fallback_row(page_num, "Respuesta JSON inválida")
        parsed["pagina"] = page_num
        return normalize_row(parsed, page_num)
    except json.JSONDecodeError as e:
        logger.error(f"JSON error página {page_num}: {e}")
        return build_fallback_row(page_num, f"Error JSON página {page_num}")
    except Exception as e:
        logger.error(f"Error página {page_num}: {e}", exc_info=True)
        return build_fallback_row(page_num, f"Error inesperado página {page_num}")


def pdf_to_page_images(pdf_bytes: bytes, dpi: int = 300) -> List[str]:
    """Convierte PDF (bytes) en lista de imágenes base64 PNG."""
    doc  = fitz.open(stream=pdf_bytes, filetype="pdf")
    zoom = dpi / 72.0
    mat  = fitz.Matrix(zoom, zoom)
    pages: List[str] = []
    for i in range(len(doc)):
        pix  = doc.load_page(i).get_pixmap(matrix=mat, alpha=False)
        pages.append(base64.b64encode(pix.tobytes("png")).decode("utf-8"))
    doc.close()
    return pages


# ─────────────────────────────────────────────
# GENERACIÓN DE EXCEL
# ─────────────────────────────────────────────

def _border(style: str, color: str) -> Border:
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


THIN_BORDER = _border("thin",   "B8CCE4")
MED_BORDER  = _border("medium", "1F3864")

HEADER_FILL  = PatternFill(fill_type="solid", fgColor="1F3864")
ALT_FILL     = PatternFill(fill_type="solid", fgColor="EBF0FA")
WHITE_FILL   = PatternFill(fill_type="solid", fgColor="FFFFFF")
TOTAL_FILL   = PatternFill(fill_type="solid", fgColor="D9E1F2")
GREEN_FILL   = PatternFill(fill_type="solid", fgColor="C6EFCE")
YELLOW_FILL  = PatternFill(fill_type="solid", fgColor="FFEB9C")
RED_FILL     = PatternFill(fill_type="solid", fgColor="FFC7CE")

WHITE_FONT   = Font(color="FFFFFF", bold=True, name="Arial", size=10)
BOLD_FONT    = Font(bold=True, name="Arial", size=10)
BASE_FONT    = Font(name="Arial", size=10)
GREEN_FONT   = Font(color="375623", bold=True, name="Arial", size=10)
YELLOW_FONT  = Font(color="9C5700", bold=True, name="Arial", size=10)
RED_FONT     = Font(color="9C0006", bold=True, name="Arial", size=10)

HEADERS = [
    "N° Factura", "Página", "Fecha Literal", "Fecha ISO",
    "Total €", "IVA %", "Base €", "Cuota IVA €",
    "Estado", "Observaciones", "Concepto", "Tipo Impositivo IVA",
]
# Anchos de columna (una por cada cabecera)
COL_WIDTHS = [14, 8, 18, 14, 12, 8, 12, 14, 22, 50, 24, 20]
NUM_COLS   = len(HEADERS)


def create_invoice_excel(rows: List[Dict[str, Any]]) -> bytes:
    """Genera el Excel y devuelve bytes."""
    wb = Workbook()

    # ── Hoja 1: Facturas ─────────────────────────────────────
    ws = wb.active
    ws.title = "Facturas"

    # Cabecera
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill      = HEADER_FILL
        c.font      = WHITE_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = MED_BORDER
    ws.row_dimensions[1].height = 32

    # Filas de datos
    for ri, row in enumerate(rows, start=2):
        data = [
            row.get("numero_factura"),
            row.get("pagina"),
            row.get("fecha_literal"),
            row.get("fecha_iso"),
            row.get("total_eur"),
            row.get("iva_pct"),
            row.get("base_eur"),
            row.get("cuota_eur"),
            row.get("estado"),
            row.get("observaciones"),
            row.get("concepto", "venta a clientes varios"),
            row.get("tipo_impositivo_iva", 10),
        ]
        ws.append(data)

        fila_fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL
        for ci in range(1, NUM_COLS + 1):
            c = ws.cell(row=ri, column=ci)
            c.border = THIN_BORDER
            c.font   = BASE_FONT
            c.fill   = fila_fill

        # Estado con color (col I = 9)
        ec = ws.cell(row=ri, column=9)
        estado = str(row.get("estado") or "")
        if estado == "COMPLETA":
            ec.fill, ec.font = GREEN_FILL,  GREEN_FONT
        elif estado == "VERIFICAR_DATOS":
            ec.fill, ec.font = YELLOW_FILL, YELLOW_FONT
        elif estado == "PENDIENTE_REVISION":
            ec.fill, ec.font = RED_FILL,    RED_FONT
        ec.alignment = Alignment(horizontal="center", vertical="center")

        # Formato numérico
        for ci in [5, 7, 8]:
            c = ws.cell(row=ri, column=ci)
            c.number_format = '#,##0.00 €'
            c.alignment     = Alignment(horizontal="right")
        ws.cell(row=ri, column=2).alignment  = Alignment(horizontal="center")
        ws.cell(row=ri, column=4).alignment  = Alignment(horizontal="center")
        ws.cell(row=ri, column=6).alignment  = Alignment(horizontal="center")
        ws.cell(row=ri, column=10).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=ri, column=11).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=ri, column=12).alignment = Alignment(horizontal="center", vertical="center")

    # Fila TOTALES
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value="TOTALES")
    for ci in range(1, NUM_COLS + 1):
        c = ws.cell(row=total_row, column=ci)
        c.fill   = TOTAL_FILL
        c.border = MED_BORDER
        c.font   = BOLD_FONT
    for ci, col_letter in [(5, "E"), (7, "G"), (8, "H")]:
        c = ws.cell(row=total_row, column=ci)
        c.value         = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
        c.number_format = '#,##0.00 €'
        c.alignment     = Alignment(horizontal="right")

    # Filtros, panel fijo, anchos
    ws.auto_filter.ref = f"A1:{get_column_letter(NUM_COLS)}1"
    ws.freeze_panes    = "A2"
    for ci, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width

    # ── Hoja 2: Resumen con fórmulas dinámicas ───────────────
    ws2 = wb.create_sheet("Resumen")
    n   = len(rows)         # número de filas de datos
    last_data_row = n + 1   # última fila de datos en hoja Facturas

    ws2.merge_cells("A1:B1")
    ws2["A1"] = "Resumen de extracción"
    ws2["A1"].fill      = HEADER_FILL
    ws2["A1"].font      = WHITE_FONT
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    # Fórmulas dinámicas referenciando hoja Facturas
    summary_data = [
        ("Total páginas procesadas",  f"=COUNTA(Facturas!B2:B{last_data_row})"),
        ("✅ Completas",               f'=COUNTIF(Facturas!I2:I{last_data_row},"COMPLETA")'),
        ("⚠️ Verificar datos",         f'=COUNTIF(Facturas!I2:I{last_data_row},"VERIFICAR_DATOS")'),
        ("❌ Pendientes revisión",      f'=COUNTIF(Facturas!I2:I{last_data_row},"PENDIENTE_REVISION")'),
        ("% Completadas",              f'=IFERROR(COUNTIF(Facturas!I2:I{last_data_row},"COMPLETA")/COUNTA(Facturas!B2:B{last_data_row}),0)'),
        ("Total facturado €",          f"=SUM(Facturas!E2:E{last_data_row})"),
        ("Base imponible €",           f"=SUM(Facturas!G2:G{last_data_row})"),
        ("Total IVA €",                f"=SUM(Facturas!H2:H{last_data_row})"),
        ("Estado general",             f'=IF(COUNTIF(Facturas!I2:I{last_data_row},"PENDIENTE_REVISION")>={max(2, n//2)},"INCOMPLETO",IF(COUNTIF(Facturas!I2:I{last_data_row},"VERIFICAR_DATOS")>0,"REVISAR","EXITOSO"))'),
    ]

    thin2 = _border("thin", "B8CCE4")
    for idx, (label, formula) in enumerate(summary_data, start=2):
        c1 = ws2.cell(row=idx, column=1, value=label)
        c2 = ws2.cell(row=idx, column=2, value=formula)
        fill = ALT_FILL if idx % 2 == 0 else WHITE_FILL
        for c in [c1, c2]:
            c.border = thin2
            c.fill   = fill
            c.font   = BASE_FONT
        c1.font = BOLD_FONT

        # Formatos especiales
        if label == "% Completadas":
            c2.number_format = '0.00%'
        elif "€" in label:
            c2.number_format = '#,##0.00 €'

    # Color estado general (fila 10)
    ws2.cell(row=10, column=2).fill = PatternFill(fill_type="solid", fgColor="EBF0FA")

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 22

    # Guardar en bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
